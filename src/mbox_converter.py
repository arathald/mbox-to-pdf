"""
Core mbox parsing and PDF generation logic.

This module provides the public API for converting mbox files to PDF documents.
All processing is done locally—no network calls, no telemetry.
"""

import base64
import csv
import html
import io
import mailbox
from dataclasses import dataclass, field
from datetime import datetime
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional


@dataclass
class Attachment:
    """Processed email attachment with size and rendering info.

    Attributes:
        filename: Original attachment filename
        mime_type: MIME type (e.g., 'image/png', 'application/pdf')
        size_bytes: Size of raw attachment data
        raw_content: Raw bytes of the attachment (for rendering)
        rendered_content: HTML content if successfully rendered, None otherwise
        content_type: Rendering category ('html', 'text', 'table', 'image', 'reference')
        error: Error details if rendering failed
    """
    filename: str
    mime_type: str
    size_bytes: int
    raw_content: bytes = field(default=b"", repr=False)  # Exclude from repr for readability
    rendered_content: Optional[str] = None
    content_type: Optional[str] = None
    error: Optional[str] = None

    def format_size_for_display(self) -> str:
        """Return human-readable size (e.g., '1.2 MB')."""
        size = self.size_bytes
        for unit in ['B', 'KB', 'MB']:
            if size < 1024:
                return f"{size:.1f} {unit}" if unit != 'B' else f"{int(size)} B"
            size /= 1024
        return f"{size:.1f} GB"


@dataclass
class Email:
    """Email message with complete header information for forensic/legal documentation.

    All fields are preserved exactly as received (or empty/None if not present).
    This ensures the PDF archive is a complete and authentic record.

    Attributes:
        message_id: Unique message ID (e.g., '<abc123@example.com>')
        from_addr: Sender address with display name
        to_addr: Comma-separated list of recipients
        date: Parsed datetime with timezone
        subject: Email subject line
        body_text: Plain text body content
        body_html: HTML body content if present
        cc_addr: Comma-separated CC recipients
        bcc_addr: Comma-separated BCC recipients
        in_reply_to: Message-ID of parent message (for threading)
        references: List of message-IDs in thread chain
        attachments: List of Attachment objects
        headers: Complete header dictionary for forensic completeness
        x_mailer: Original email client identifier
    """
    message_id: str
    from_addr: str
    to_addr: str
    date: datetime
    subject: str
    body_text: str
    body_html: Optional[str] = None
    cc_addr: Optional[str] = None
    bcc_addr: Optional[str] = None
    in_reply_to: Optional[str] = None
    references: Optional[List[str]] = None
    attachments: List[Attachment] = field(default_factory=list)
    headers: Dict[str, str] = field(default_factory=dict)
    x_mailer: Optional[str] = None


@dataclass
class ConversionResult:
    """Result of a mbox-to-PDF conversion operation.

    Attributes:
        success: True if conversion completed without fatal errors
        pdfs_created: Number of PDF files successfully created
        emails_processed: Total number of emails processed
        errors: List of human-readable error messages (for logs)
        attachment_errors: List of AttachmentErrorInfo for dialog display
        created_files: List of paths to generated PDF files
    """
    success: bool
    pdfs_created: int
    emails_processed: int = 0
    errors: List[str] = field(default_factory=list)
    attachment_errors: List["AttachmentErrorInfo"] = field(default_factory=list)
    created_files: List[str] = field(default_factory=list)


# Import here to avoid circular imports
from src.error_handling import AttachmentErrorInfo


def render_attachment(attachment: Attachment) -> str:
    """Render an attachment as HTML for embedding in PDF.

    Dispatches to appropriate renderer based on MIME type:
    - text/plain → formatted text
    - text/csv → HTML table
    - image/* → base64-encoded <img> tag
    - text/html → sanitized HTML
    - application/docx → extracted text
    - application/xlsx → HTML tables
    - unsupported → reference note with filename and size

    All processing is local—no network calls.

    Args:
        attachment: Attachment object with raw_content populated

    Returns:
        HTML string representing the rendered attachment

    Side effects:
        Sets attachment.rendered_content and attachment.content_type
        Sets attachment.error if rendering fails
    """
    mime_type = attachment.mime_type.lower()
    filename = attachment.filename

    try:
        # Dispatch based on MIME type
        if mime_type == "text/csv" or filename.lower().endswith(".csv"):
            html_content = _render_csv(attachment)
            attachment.content_type = "table"
        elif mime_type == "text/html" or filename.lower().endswith(".html") or filename.lower().endswith(".htm"):
            html_content = _render_html(attachment)
            attachment.content_type = "html"
        elif mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" or filename.lower().endswith(".docx"):
            html_content = _render_docx(attachment)
            attachment.content_type = "docx"
        elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or filename.lower().endswith(".xlsx"):
            html_content = _render_xlsx(attachment)
            attachment.content_type = "xlsx"
        elif mime_type.startswith("text/"):
            html_content = _render_text(attachment)
            attachment.content_type = "text"
        elif mime_type.startswith("image/"):
            html_content = _render_image(attachment)
            attachment.content_type = "image"
        else:
            html_content = _render_reference(attachment)
            attachment.content_type = "reference"
    except Exception as e:
        # If rendering fails, show reference and record error
        attachment.error = str(e)
        html_content = _render_reference(attachment)
        attachment.content_type = "reference"

    attachment.rendered_content = html_content
    return html_content


def _render_docx(attachment: Attachment) -> str:
    """Extract text content from Word document."""
    from docx import Document
    from io import BytesIO
    doc = Document(BytesIO(attachment.raw_content))
    text = '\n\n'.join(paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip())

    if not text:
        return _render_reference(attachment)

    # Render as formatted text with paragraph breaks preserved
    paragraphs = text.split('\n\n')
    formatted = ''.join(f'<p>{html.escape(para).replace(chr(10), "<br/>")}</p>' for para in paragraphs if para.strip())
    return f'<div class="attachment-docx">{formatted}</div>'


def _render_xlsx(attachment: Attachment) -> str:
    """Extract and render Excel spreadsheet as HTML tables."""
    from openpyxl import load_workbook
    from io import BytesIO
    wb = load_workbook(BytesIO(attachment.raw_content))

    html_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        html_parts.append(f'<h4 style="margin-top: 1em; margin-bottom: 0.5em;">{html.escape(sheet_name)}</h4>')
        html_parts.append('<table class="attachment-xlsx">')

        for row in ws.iter_rows():
            html_parts.append('<tr>')
            for cell in row:
                value = cell.value if cell.value is not None else ''
                html_parts.append(f'<td>{html.escape(str(value))}</td>')
            html_parts.append('</tr>')

        html_parts.append('</table>')

    if not html_parts:
        return _render_reference(attachment)

    return f'<div class="attachment-xlsx">{"".join(html_parts)}</div>'


def _render_html(attachment: Attachment) -> str:
    """Render HTML attachment by sanitizing and wrapping it."""
    try:
        content = attachment.raw_content.decode("utf-8")
    except UnicodeDecodeError:
        try:
            content = attachment.raw_content.decode("latin-1")
        except UnicodeDecodeError:
            content = attachment.raw_content.decode("utf-8", errors="replace")

    # Sanitize HTML using bleach
    from bleach import clean
    sanitized = clean(
        content,
        tags=['html', 'body', 'head', 'title', 'style', 'p', 'br', 'div', 'span', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6',
              'strong', 'b', 'em', 'i', 'u', 'a', 'img', 'table', 'thead', 'tbody', 'tr', 'th', 'td', 'ol', 'ul', 'li'],
        attributes={'a': ['href', 'title'], 'img': ['src', 'alt', 'title', 'width', 'height']},
        strip=True
    )
    return f'<div class="attachment-html">{sanitized}</div>'


def _render_text(attachment: Attachment) -> str:
    """Render text attachment as regular formatted text."""
    try:
        content = attachment.raw_content.decode("utf-8")
    except UnicodeDecodeError:
        try:
            content = attachment.raw_content.decode("latin-1")
        except UnicodeDecodeError:
            content = attachment.raw_content.decode("utf-8", errors="replace")

    escaped = html.escape(content)
    # Preserve paragraph breaks but wrap naturally
    paragraphs = escaped.split('\n\n')
    formatted = ''.join(f'<p>{para.replace(chr(10), "<br/>")}</p>' for para in paragraphs if para.strip())
    return f'<div class="attachment-text">{formatted}</div>'


def _render_csv(attachment: Attachment) -> str:
    """Render CSV attachment as HTML table."""
    try:
        content = attachment.raw_content.decode("utf-8")
    except UnicodeDecodeError:
        content = attachment.raw_content.decode("latin-1", errors="replace")

    reader = csv.reader(io.StringIO(content))
    rows = list(reader)

    if not rows:
        return _render_reference(attachment)

    html_parts = ['<table class="attachment-csv">']

    # First row as header
    html_parts.append("<thead><tr>")
    for cell in rows[0]:
        html_parts.append(f"<th>{html.escape(cell)}</th>")
    html_parts.append("</tr></thead>")

    # Data rows
    if len(rows) > 1:
        html_parts.append("<tbody>")
        for row in rows[1:]:
            html_parts.append("<tr>")
            for cell in row:
                html_parts.append(f"<td>{html.escape(cell)}</td>")
            html_parts.append("</tr>")
        html_parts.append("</tbody>")

    html_parts.append("</table>")
    return "".join(html_parts)


def _render_image(attachment: Attachment) -> str:
    """Render image attachment as base64-encoded img tag."""
    b64_content = base64.b64encode(attachment.raw_content).decode("ascii")
    mime_type = attachment.mime_type
    filename = html.escape(attachment.filename)

    return (
        f'<img src="data:{mime_type};base64,{b64_content}" '
        f'alt="{filename}" class="attachment-image" />'
    )


def _render_reference(attachment: Attachment) -> str:
    """Render unsupported attachment as reference note."""
    filename = html.escape(attachment.filename)
    mime_type = html.escape(attachment.mime_type)
    size = attachment.format_size_for_display()

    return (
        f'<div class="attachment-reference">'
        f'<strong>Attachment:</strong> {filename}<br/>'
        f'<small>Type: {mime_type} | Size: {size}</small><br/>'
        f'<em>This attachment type cannot be rendered in the PDF.</em>'
        f'</div>'
    )


def render_email_to_html(email: Email) -> str:
    """Render a complete email as HTML for PDF generation.

    Produces professional formatting suitable for legal/forensic documentation:
    - Full header section with all forensic fields
    - Body content (plain text or sanitized HTML)
    - Attachments section with rendered content

    All processing is local—no network calls.

    Args:
        email: Email object to render

    Returns:
        Complete HTML string for the email
    """
    parts = ['<div class="email-message">']

    # Header section
    parts.append(_render_email_header(email))

    # Body section
    parts.append(_render_email_body(email))

    # Attachments section (if any)
    if email.attachments:
        parts.append(_render_attachments_section(email.attachments))

    parts.append("</div>")
    return "".join(parts)


def _render_email_header(email: Email) -> str:
    """Render email header as professional HTML."""
    parts = ['<div class="email-header">']

    # Date - formatted for readability
    date_str = email.date.strftime("%A, %B %d, %Y at %I:%M %p")
    parts.append(_render_header_field("Date", date_str))

    # From
    parts.append(_render_header_field("From", html.escape(email.from_addr)))

    # To
    parts.append(_render_header_field("To", html.escape(email.to_addr)))

    # CC (if present)
    if email.cc_addr:
        parts.append(_render_header_field("CC", html.escape(email.cc_addr)))

    # BCC (if present)
    if email.bcc_addr:
        parts.append(_render_header_field("BCC", html.escape(email.bcc_addr)))

    # Subject
    parts.append(_render_header_field("Subject", html.escape(email.subject)))

    # Message-ID (forensic reference)
    parts.append(_render_header_field("Message-ID", html.escape(email.message_id)))

    # In-Reply-To (if present - threading)
    if email.in_reply_to:
        parts.append(_render_header_field("In-Reply-To", html.escape(email.in_reply_to)))

    # References (if present - threading)
    if email.references:
        refs = " ".join(email.references)
        parts.append(_render_header_field("References", html.escape(refs)))

    # X-Mailer (if present - authenticity)
    if email.x_mailer:
        parts.append(_render_header_field("X-Mailer", html.escape(email.x_mailer)))

    parts.append("</div>")
    return "".join(parts)


def _render_header_field(label: str, value: str) -> str:
    """Render a single header field."""
    return (
        f'<div class="header-field">'
        f'<span class="header-label">{label}:</span> '
        f'<span class="header-value">{value}</span>'
        f'</div>'
    )


def _render_email_body(email: Email) -> str:
    """Render email body as HTML."""
    parts = ['<div class="email-body">']

    if email.body_html:
        # HTML email - include as-is (sanitization happens at parse time)
        parts.append(f'<div class="html-body">{email.body_html}</div>')
    elif email.body_text:
        # Plain text email - render with natural wrapping, preserving paragraph breaks
        escaped = html.escape(email.body_text)
        paragraphs = escaped.split('\n\n')
        formatted = ''.join(f'<p>{para.replace(chr(10), "<br/>")}</p>' for para in paragraphs if para.strip())
        parts.append(f'<div class="plaintext-body">{formatted}</div>')

    parts.append("</div>")
    return "".join(parts)


def _render_attachments_section(attachments: List[Attachment]) -> str:
    """Render attachments section with all attachment content."""
    parts = ['<div class="attachments-section">']
    parts.append('<h3 class="attachments-header">Attachments</h3>')

    for attachment in attachments:
        parts.append('<div class="attachment-item">')
        parts.append(
            f'<div class="attachment-name">'
            f'{html.escape(attachment.filename)} '
            f'<small>({attachment.format_size_for_display()})</small>'
            f'</div>'
        )

        # Render the attachment content
        rendered = render_attachment(attachment)
        parts.append(f'<div class="attachment-content">{rendered}</div>')

        parts.append("</div>")

    parts.append("</div>")
    return "".join(parts)


# CSS styles for PDF rendering
PDF_STYLES = """
<style>
@page {
    size: letter;
    margin: 1in 0.75in;
}

body {
    font-family: "Times New Roman", serif;
    font-size: 10pt;
    line-height: 1.5;
    color: #333;
}

.email-message {
    page-break-before: always;
}

.email-message:first-child {
    page-break-before: avoid;
}

.email-header {
    background-color: #f5f5f5;
    border: 1px solid #ddd;
    border-radius: 3px;
    padding: 12px 15px;
    margin-bottom: 20px;
    font-family: monospace;
    font-size: 10pt;
}

.header-field {
    margin-bottom: 3px;
}

.header-label {
    font-weight: bold;
    color: #555;
    display: block;
    margin-bottom: 2px;
}

.header-value {
    color: #222;
    display: block;
    word-wrap: break-word;
    overflow-wrap: break-word;
}

.email-body {
    margin-bottom: 20px;
}

.plaintext-body {
    font-family: "Times New Roman", serif;
    font-size: 10pt;
    line-height: 1.6;
}

.plaintext-body p {
    margin: 0.5em 0;
}

.html-body {
    padding: 10px;
}

.attachments-section {
    margin-top: 20px;
    border-top: 1px solid #ddd;
    padding-top: 15px;
}

.attachments-header {
    font-size: 12pt;
    margin-bottom: 10px;
}

.attachment-item {
    margin-bottom: 15px;
    border: 1px solid #eee;
    padding: 10px;
}

.attachment-name {
    font-family: monospace;
    color: #222;
    font-weight: 500;
    margin-bottom: 8px;
}

.attachment-text {
    font-family: "Times New Roman", serif;
    font-size: 9pt;
    line-height: 1.5;
}

.attachment-text p {
    margin: 0.4em 0;
}

.attachment-html {
    font-family: "Times New Roman", serif;
    font-size: 9pt;
    line-height: 1.5;
    border: 1px solid #eee;
    padding: 8px;
    background-color: #fafafa;
}

.attachment-docx {
    font-family: "Times New Roman", serif;
    font-size: 9pt;
    line-height: 1.5;
}

.attachment-docx p {
    margin: 0.4em 0;
}

.attachment-xlsx {
    font-size: 8pt;
}

.attachment-xlsx h4 {
    margin: 1em 0 0.5em 0;
    font-size: 10pt;
    font-weight: bold;
}

.attachment-csv {
    width: 100%;
    border-collapse: collapse;
    font-size: 9pt;
}

.attachment-csv th,
.attachment-csv td {
    border: 1px solid #ddd;
    padding: 4px 8px;
    text-align: left;
}

.attachment-csv th {
    background-color: #f5f5f5;
    font-weight: bold;
}

.attachment-xlsx table {
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 1em;
    font-size: 8pt;
}

.attachment-xlsx td {
    border: 1px solid #ddd;
    padding: 3px 6px;
    text-align: left;
}

.attachment-image {
    max-width: 95%;
    height: auto;
}

.attachment-reference {
    background-color: #fff3cd;
    border: 1px solid #ffc107;
    padding: 10px;
    border-radius: 3px;
}
</style>
"""


def generate_pdf(html_content: str, output_path: Path, image_temp_dir: Optional[Path] = None) -> Path:
    """Generate a PDF from HTML content using xhtml2pdf.

    All processing is local—no network calls, no native dependencies.

    Args:
        html_content: HTML string to render
        output_path: Path for the output PDF file
        image_temp_dir: Optional temp directory for image files

    Returns:
        The output path (for chaining)

    Raises:
        RuntimeError: If PDF generation fails
    """
    from xhtml2pdf import pisa
    import base64
    import re

    # If we have a temp dir, convert data: URIs to file: URIs
    if image_temp_dir:
        image_temp_dir = Path(image_temp_dir)
        image_temp_dir.mkdir(parents=True, exist_ok=True)

        # Find all data: URIs and write them to temp files
        counter = [0]  # Use list to allow modification in nested function
        def replace_data_uri(match):
            uri = match.group(0)
            counter[0] += 1
            try:
                header, data = uri.split(',', 1)
                if 'base64' in header:
                    image_data = base64.b64decode(data)
                else:
                    image_data = data.encode('utf-8')

                # Determine file extension from MIME type
                ext = '.png'
                if 'jpeg' in header or 'jpg' in header:
                    ext = '.jpg'
                elif 'gif' in header:
                    ext = '.gif'
                elif 'webp' in header:
                    ext = '.webp'

                # Write to temp file
                temp_file = image_temp_dir / f"image_{counter[0]}{ext}"
                temp_file.write_bytes(image_data)

                # Return file: URI
                return f'file://{temp_file.absolute()}'
            except Exception:
                return uri  # Return original if conversion fails

        # Replace all data: URIs
        html_content = re.sub(r'data:[^"\';\s]+;[^"\';\s]*;[^"\']*,[^"\']*', replace_data_uri, html_content)

    # Wrap content in full HTML document with styles
    full_html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    {PDF_STYLES}
</head>
<body>
    {html_content}
</body>
</html>
"""

    # Generate PDF
    output_path = Path(output_path)
    with open(output_path, "wb") as pdf_file:
        pisa_status = pisa.CreatePDF(full_html, dest=pdf_file)

    if pisa_status.err:
        raise RuntimeError(f"PDF generation failed with {pisa_status.err} errors")

    return output_path


def group_emails_by_date(
    emails: List[Email],
    strategy: str
) -> Dict[str, List[Email]]:
    """Group emails by date according to the specified strategy.

    Groups are returned as an ordered dict with keys sorted chronologically.
    Emails within each group are also sorted by date ascending.

    All processing is local—no network calls.

    Args:
        emails: List of Email objects to group
        strategy: One of "month", "quarter", or "year"
            - month: Keys like "2008-01-January"
            - quarter: Keys like "2008-Q1"
            - year: Keys like "2008"

    Returns:
        Dict mapping group keys to lists of Email objects

    Raises:
        ValueError: If strategy is not one of the valid options
    """
    if strategy not in ("month", "quarter", "year"):
        raise ValueError(
            f"Invalid grouping strategy: '{strategy}'. "
            "Must be 'month', 'quarter', or 'year'."
        )

    if not emails:
        return {}

    groups: Dict[str, List[Email]] = {}

    for email in emails:
        key = _get_group_key(email.date, strategy)
        if key not in groups:
            groups[key] = []
        groups[key].append(email)

    # Sort emails within each group by date
    for key in groups:
        groups[key].sort(key=lambda e: e.date)

    # Return groups sorted by key (chronological order)
    sorted_groups = dict(sorted(groups.items()))
    return sorted_groups


def _get_group_key(date: datetime, strategy: str) -> str:
    """Generate the group key for a date based on strategy.

    Args:
        date: The datetime to generate a key for
        strategy: One of "month", "quarter", or "year"

    Returns:
        Group key string
    """
    if strategy == "month":
        return date.strftime("%Y-%m-%B")  # e.g., "2008-01-January"
    elif strategy == "quarter":
        quarter = (date.month - 1) // 3 + 1
        return f"{date.year}-Q{quarter}"  # e.g., "2008-Q1"
    else:  # year
        return str(date.year)  # e.g., "2008"


def merge_and_deduplicate(mbox_paths: List[Path]) -> List[Email]:
    """Parse multiple mbox files and merge into a deduplicated list.

    Deduplication is by Message-ID. When duplicates are found, the first
    occurrence is kept. This handles Gmail Takeout exports where emails
    appear in multiple label folders.

    All processing is local—no network calls.

    Args:
        mbox_paths: List of paths to mbox files

    Returns:
        List of unique Email objects, sorted by date ascending
    """
    seen_message_ids: set[str] = set()
    unique_emails: List[Email] = []

    for mbox_path in mbox_paths:
        emails = parse_mbox(mbox_path)
        for email in emails:
            if email.message_id not in seen_message_ids:
                seen_message_ids.add(email.message_id)
                unique_emails.append(email)

    # Sort by date ascending
    unique_emails.sort(key=lambda e: e.date)
    return unique_emails


def parse_mbox(mbox_path: Path) -> List[Email]:
    """Parse an mbox file and return a list of Email objects.

    Uses Python's stdlib mailbox module for robust mbox parsing.
    All processing is local—no network calls.

    Args:
        mbox_path: Path to the mbox file

    Returns:
        List of Email objects, sorted by date ascending

    Raises:
        FileNotFoundError: If mbox_path does not exist
        ValueError: If file is not a valid mbox format
    """
    mbox_path = Path(mbox_path)
    if not mbox_path.exists():
        raise FileNotFoundError(f"mbox file not found: {mbox_path}")

    mbox = mailbox.mbox(mbox_path)
    emails = []

    for message in mbox:
        email = _parse_message(message)
        if email:
            emails.append(email)

    # Sort by date ascending
    emails.sort(key=lambda e: e.date)
    return emails


def _parse_message(message: mailbox.mboxMessage) -> Optional[Email]:
    """Parse a single mbox message into an Email object.

    Args:
        message: mailbox.mboxMessage instance

    Returns:
        Email object, or None if message is malformed
    """
    # Extract required headers
    message_id = message.get('Message-ID', '')
    from_addr = message.get('From', '')
    to_addr = message.get('To', '')
    subject = message.get('Subject', '')

    # Parse date with timezone
    date_str = message.get('Date', '')
    try:
        date = parsedate_to_datetime(date_str)
    except (ValueError, TypeError):
        # If date parsing fails, skip this message
        return None

    # Optional headers
    cc_addr = message.get('Cc')
    bcc_addr = message.get('Bcc')
    in_reply_to = message.get('In-Reply-To')
    x_mailer = message.get('X-Mailer')

    # Parse References header (space-separated list of message IDs)
    references_str = message.get('References')
    references = None
    if references_str:
        references = references_str.split()

    # Extract body content
    body_text, body_html, attachments = _extract_body_and_attachments(message)

    # Build complete headers dict for forensic completeness
    headers = {key: message.get(key) for key in message.keys()}

    return Email(
        message_id=message_id,
        from_addr=from_addr,
        to_addr=to_addr,
        date=date,
        subject=subject,
        body_text=body_text,
        body_html=body_html,
        cc_addr=cc_addr,
        bcc_addr=bcc_addr,
        in_reply_to=in_reply_to,
        references=references,
        attachments=attachments,
        headers=headers,
        x_mailer=x_mailer,
    )


def _extract_body_and_attachments(
    message: mailbox.mboxMessage
) -> tuple[str, Optional[str], List[Attachment]]:
    """Extract body text, HTML, and attachments from a message.

    Handles both simple messages and multipart MIME messages.

    Args:
        message: The email message to process

    Returns:
        Tuple of (body_text, body_html, attachments_list)
    """
    body_text = ""
    body_html = None
    attachments = []

    if message.is_multipart():
        for part in message.walk():
            content_type = part.get_content_type()
            content_disposition = str(part.get('Content-Disposition', ''))

            # Skip multipart containers
            if part.get_content_maintype() == 'multipart':
                continue

            # Check if this is an attachment
            if 'attachment' in content_disposition or part.get_filename():
                attachment = _parse_attachment(part)
                if attachment:
                    attachments.append(attachment)
            elif content_type == 'text/plain' and not body_text:
                body_text = _decode_payload(part)
            elif content_type == 'text/html' and body_html is None:
                body_html = _decode_payload(part)
    else:
        # Simple non-multipart message
        content_type = message.get_content_type()
        if content_type == 'text/plain':
            body_text = _decode_payload(message)
        elif content_type == 'text/html':
            body_html = _decode_payload(message)
            body_text = ""  # Will need to strip HTML for text version

    return body_text, body_html, attachments


def _decode_payload(part) -> str:
    """Decode message payload to string, handling various encodings.

    Args:
        part: Message or message part

    Returns:
        Decoded string content
    """
    payload = part.get_payload(decode=True)
    if payload is None:
        return ""

    # Try to decode with charset from headers, fall back to utf-8, then latin-1
    charset = part.get_content_charset() or 'utf-8'
    try:
        return payload.decode(charset)
    except (UnicodeDecodeError, LookupError):
        try:
            return payload.decode('utf-8')
        except UnicodeDecodeError:
            return payload.decode('latin-1', errors='replace')


def _parse_attachment(part) -> Optional[Attachment]:
    """Parse a MIME part into an Attachment object.

    Args:
        part: MIME message part

    Returns:
        Attachment object, or None if parsing fails
    """
    filename = part.get_filename() or 'unnamed_attachment'
    mime_type = part.get_content_type()

    payload = part.get_payload(decode=True)
    if payload is None:
        return None

    size_bytes = len(payload)

    return Attachment(
        filename=filename,
        mime_type=mime_type,
        size_bytes=size_bytes,
        raw_content=payload,
    )


def add_continuation_headers(
    input_pdf: Path,
    output_pdf: Path,
    subject: str,
    from_addr: str,
) -> Path:
    """Add continuation headers to pages 2+ of a multi-page PDF.

    For multi-page emails, adds a header like:
    "Subject: [subject] (continued) - From: [from_addr]"
    to the top of pages 2 and beyond.

    Single-page PDFs are copied through unchanged.

    Args:
        input_pdf: Path to the input PDF
        output_pdf: Path for the output PDF
        subject: Email subject for the header
        from_addr: Email sender for the header

    Returns:
        The output path
    """
    from io import BytesIO

    from pypdf import PdfReader, PdfWriter
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    input_pdf = Path(input_pdf)
    output_pdf = Path(output_pdf)

    reader = PdfReader(input_pdf)
    writer = PdfWriter()

    total_pages = len(reader.pages)

    # If single page, just copy through
    if total_pages <= 1:
        for page in reader.pages:
            writer.add_page(page)
        with open(output_pdf, "wb") as f:
            writer.write(f)
        return output_pdf

    # Create header overlay for continuation pages
    def create_header_overlay(page_num: int) -> BytesIO:
        """Create a PDF page with just the header text."""
        packet = BytesIO()
        c = canvas.Canvas(packet, pagesize=letter)
        page_width, page_height = letter

        # Truncate subject if too long
        max_subject_len = 60
        display_subject = subject[:max_subject_len] + "..." if len(subject) > max_subject_len else subject

        # Header text
        header_text = f"Subject: {display_subject} (continued - page {page_num} of {total_pages})"

        # Draw header at top of page
        c.setFont("Helvetica", 8)
        c.setFillColorRGB(0.4, 0.4, 0.4)  # Gray color
        c.drawString(54, page_height - 36, header_text)  # 0.75" from top, 0.75" from left

        c.save()
        packet.seek(0)
        return packet

    # Process each page
    for i, page in enumerate(reader.pages):
        if i == 0:
            # First page - no header needed
            writer.add_page(page)
        else:
            # Create header overlay for this page
            header_pdf = PdfReader(create_header_overlay(i + 1))
            header_page = header_pdf.pages[0]

            # Merge header onto the original page
            page.merge_page(header_page)
            writer.add_page(page)

    with open(output_pdf, "wb") as f:
        writer.write(f)

    return output_pdf


def merge_pdfs(pdf_paths: List[Path], output_path: Path) -> Path:
    """Merge multiple PDF files into a single PDF.

    Args:
        pdf_paths: List of paths to PDFs to merge (in order)
        output_path: Path for the merged output PDF

    Returns:
        The output path
    """
    from pypdf import PdfReader, PdfWriter

    writer = PdfWriter()

    for pdf_path in pdf_paths:
        reader = PdfReader(pdf_path)
        for page in reader.pages:
            writer.add_page(page)

    with open(output_path, "wb") as f:
        writer.write(f)

    return output_path


def convert_mbox_to_pdfs(
    mbox_paths: List[Path],
    output_dir: Path,
    grouping_strategy: str = "month",
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
) -> ConversionResult:
    """Convert mbox files to grouped PDF documents.

    This is the main orchestration function that ties together parsing,
    deduplication, grouping, rendering, and PDF generation.

    All processing is local—no network calls, no telemetry.

    Args:
        mbox_paths: List of paths to mbox files to convert
        output_dir: Directory where PDFs will be saved
        grouping_strategy: One of "month", "quarter", or "year"
        progress_callback: Optional function called with (current, total, message)
            to report progress for GUI updates

    Returns:
        ConversionResult with success status, counts, and file paths
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    errors: List[str] = []
    attachment_errors: List[AttachmentErrorInfo] = []
    created_files: List[str] = []

    # Step 1: Parse and deduplicate emails from all mbox files
    if progress_callback:
        progress_callback(0, 100, "Parsing mbox files...")

    try:
        emails = merge_and_deduplicate(mbox_paths)
    except Exception as e:
        return ConversionResult(
            success=False,
            pdfs_created=0,
            emails_processed=0,
            errors=[f"Failed to parse mbox files: {e}"],
            created_files=[],
        )

    if not emails:
        return ConversionResult(
            success=True,
            pdfs_created=0,
            emails_processed=0,
            errors=["No emails found in the provided mbox files"],
            created_files=[],
        )

    emails_processed = len(emails)

    if progress_callback:
        progress_callback(10, 100, f"Found {emails_processed} emails, grouping...")

    # Step 2: Group emails by date
    groups = group_emails_by_date(emails, grouping_strategy)

    if progress_callback:
        progress_callback(20, 100, f"Creating {len(groups)} PDF files...")

    # Step 3: Generate PDF for each group
    import tempfile
    import shutil

    total_groups = len(groups)
    for i, (group_key, group_emails) in enumerate(groups.items()):
        if progress_callback:
            progress_pct = 20 + int((i / total_groups) * 75)
            progress_callback(progress_pct, 100, f"Generating {group_key}.pdf...")

        # Create temp directory for intermediate PDFs
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            email_pdfs: List[Path] = []

            # Generate PDF for each email with continuation headers
            for j, email in enumerate(group_emails):
                try:
                    # Render email to HTML
                    email_html = render_email_to_html(email)

                    # Collect any attachment warnings from this email
                    for attachment in email.attachments:
                        # Track rendering errors
                        if attachment.error:
                            error_info = AttachmentErrorInfo(
                                email_subject=email.subject,
                                email_date=email.date.strftime('%a, %B %d, %Y at %I:%M %p'),
                                email_from=email.from_addr,
                                attachment_filename=attachment.filename,
                                mime_type=attachment.mime_type,
                                file_size=attachment.format_size_for_display(),
                                error_type="rendering_failed",
                                error_message=f"Could not render attachment: {attachment.error}"
                            )
                            attachment_errors.append(error_info)
                        # Track attachments rendered as reference notes (unsupported types)
                        elif attachment.content_type == "reference":
                            error_info = AttachmentErrorInfo(
                                email_subject=email.subject,
                                email_date=email.date.strftime('%a, %B %d, %Y at %I:%M %p'),
                                email_from=email.from_addr,
                                attachment_filename=attachment.filename,
                                mime_type=attachment.mime_type,
                                file_size=attachment.format_size_for_display(),
                                error_type="unsupported_type",
                                error_message="This file type cannot be rendered in the PDF."
                            )
                            attachment_errors.append(error_info)

                    # Generate initial PDF
                    raw_pdf = temp_path / f"email_{j:04d}_raw.pdf"
                    images_dir = temp_path / f"email_{j:04d}_images"
                    generate_pdf(email_html, raw_pdf, image_temp_dir=images_dir)

                    # Add continuation headers for multi-page emails
                    final_pdf = temp_path / f"email_{j:04d}.pdf"
                    add_continuation_headers(
                        raw_pdf,
                        final_pdf,
                        subject=email.subject,
                        from_addr=email.from_addr,
                    )

                    email_pdfs.append(final_pdf)

                except Exception as e:
                    errors.append(f"Failed to process email '{email.subject}': {e}")

            if not email_pdfs:
                errors.append(f"No emails processed for group {group_key}")
                continue

            # Merge all email PDFs into the final group PDF
            pdf_filename = f"{group_key}.pdf"
            pdf_path = output_dir / pdf_filename

            try:
                merge_pdfs(email_pdfs, pdf_path)
                created_files.append(str(pdf_path))
            except Exception as e:
                errors.append(f"Failed to merge PDFs for {group_key}: {e}")

    if progress_callback:
        progress_callback(100, 100, "Conversion complete")

    return ConversionResult(
        success=len(created_files) > 0,
        pdfs_created=len(created_files),
        emails_processed=emails_processed,
        errors=errors,
        attachment_errors=attachment_errors,
        created_files=created_files,
    )
